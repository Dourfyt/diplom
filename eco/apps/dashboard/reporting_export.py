"""
Экспорт сводного отчёта KPI (страница «Отчётность») в Excel и PDF.
"""

from __future__ import annotations

from datetime import date
from decimal import Decimal
from io import BytesIO

from django.http import HttpResponse
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Font
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle
from reportlab.lib.units import mm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from apps.integrations.remote_data import _parse_optional_date, get_remote_service
from apps.operations.pdf_fonts import ensure_pdf_cyrillic_fonts


def _fmt(value, max_decimals: int = 3) -> str:
    d = Decimal(value) if not isinstance(value, Decimal) else value
    quant = Decimal("1").scaleb(-max_decimals)
    text = format(d.quantize(quant), "f")
    if "." in text:
        text = text.rstrip("0").rstrip(".")
    return text or "0"


def load_reporting_bundle(request) -> tuple[dict, date, date, str]:
    today = date.today()
    default_from = today.replace(month=1, day=1)
    date_from = _parse_optional_date(request.GET.get("date_from")) or default_from
    date_to = _parse_optional_date(request.GET.get("date_to")) or today
    if date_from > date_to:
        date_from, date_to = date_to, date_from
    organization_id = request.GET.get("organization", "").strip()
    bundle = get_remote_service().reporting_dashboard_bundle(
        date_from=date_from,
        date_to=date_to,
        organization_id=organization_id or None,
    )
    period = f"{date_from.strftime('%d.%m.%Y')} — {date_to.strftime('%d.%m.%Y')}"
    return bundle, date_from, date_to, period


def build_reporting_excel(bundle: dict, *, period: str, organization_name: str) -> bytes:
    wb = Workbook()

    ws = wb.active
    ws.title = "KPI"
    header_font = Font(bold=True)
    ws["A1"] = "Сводный отчёт KPI"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = f"Период: {period}"
    ws["A3"] = f"Организация: {organization_name or 'Все'}"
    ws["A4"] = f"Сформирован: {timezone.localtime().strftime('%d.%m.%Y %H:%M')}"

    rows = [
        ("Партий на платформе", bundle["remote_total_batches"]),
        ("Обработано партий", bundle["remote_batches_processed"]),
        ("Объём на складе, т", _fmt(bundle["remote_total_volume"])),
        ("Партий за период", bundle["batches_total"]),
        ("Объём партий за период, т", _fmt(bundle["batch_volume"])),
        ("Операций за период", bundle["movements_total"]),
        ("Объём операций за период, т", _fmt(bundle["operation_volume"])),
        ("Измерений за период", bundle["measurements_total"]),
        ("Превышений за период", bundle["exceed_count"]),
        (
            "Выполнение плана, %",
            bundle.get("remote_plan_completion")
            if bundle.get("remote_plan_completion") is not None
            else "—",
        ),
        (
            "Средний класс опасности",
            bundle.get("remote_avg_hazard")
            if bundle.get("remote_avg_hazard") is not None
            else "—",
        ),
    ]
    row = 6
    for title, value in rows:
        ws.cell(row=row, column=1, value=title).font = header_font
        ws.cell(row=row, column=2, value=value)
        row += 1

    ws2 = wb.create_sheet("Классы опасности")
    for col, title in enumerate(("Класс", "Партий", "Объём, т"), start=1):
        ws2.cell(row=1, column=col, value=title).font = header_font
    for i, item in enumerate(bundle["hazard_rows"], start=2):
        ws2.cell(row=i, column=1, value=item["hazard_label"])
        ws2.cell(row=i, column=2, value=item["batch_count"])
        ws2.cell(row=i, column=3, value=float(item["volume_tons"]))

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def build_reporting_pdf(bundle: dict, *, period: str, organization_name: str) -> bytes:
    font_regular, font_bold = ensure_pdf_cyrillic_fonts()
    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        rightMargin=16 * mm,
        leftMargin=16 * mm,
        topMargin=14 * mm,
        bottomMargin=14 * mm,
        title="Отчёт KPI",
    )
    body = ParagraphStyle("Body", fontName=font_regular, fontSize=10)
    title = ParagraphStyle(
        "Title", fontName=font_regular, fontSize=15, spaceAfter=8, alignment=1
    )
    sub = ParagraphStyle(
        "Sub", fontName=font_regular, fontSize=9, textColor=colors.grey, alignment=1
    )

    story = [
        Paragraph("Сводный отчёт KPI", title),
        Paragraph(f"Период: {period}", sub),
        Paragraph(f"Организация: {organization_name or 'Все'}", sub),
        Paragraph(
            f"Дата формирования: {timezone.localtime().strftime('%d.%m.%Y %H:%M')}",
            sub,
        ),
        Spacer(1, 12),
    ]

    kpi_data = [
        ["Показатель", "Значение"],
        ["Партий на платформе", str(bundle["remote_total_batches"])],
        ["Обработано партий", str(bundle["remote_batches_processed"])],
        ["Объём на складе, т", _fmt(bundle["remote_total_volume"])],
        ["Партий за период", str(bundle["batches_total"])],
        ["Объём партий за период, т", _fmt(bundle["batch_volume"])],
        ["Операций за период", str(bundle["movements_total"])],
        ["Объём операций, т", _fmt(bundle["operation_volume"])],
        ["Измерений за период", str(bundle["measurements_total"])],
        ["Превышений", str(bundle["exceed_count"])],
    ]
    plan = bundle.get("remote_plan_completion")
    kpi_data.append(
        ["Выполнение плана, %", _fmt(plan, max_decimals=1) if plan is not None else "—"]
    )
    tbl = Table(kpi_data, colWidths=[doc.width * 0.62, doc.width * 0.38])
    tbl.setStyle(
        TableStyle(
            [
                ("FONTNAME", (0, 0), (-1, 0), font_bold),
                ("FONTNAME", (0, 1), (-1, -1), font_regular),
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#6c757d")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("GRID", (0, 0), (-1, -1), 0.4, colors.lightgrey),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8f9fa")]),
            ]
        )
    )
    story.append(tbl)
    story.append(Spacer(1, 14))
    story.append(Paragraph("Объём по классам опасности", body))
    story.append(Spacer(1, 6))

    hazard_data = [["Класс", "Партий", "Объём, т"]]
    for item in bundle["hazard_rows"]:
        hazard_data.append(
            [
                item["hazard_label"],
                str(item["batch_count"]),
                _fmt(item["volume_tons"]),
            ]
        )
    if len(hazard_data) == 1:
        hazard_data.append(["—", "0", "0"])

    htbl = Table(hazard_data, colWidths=[doc.width * 0.2, doc.width * 0.25, doc.width * 0.55])
    htbl.setStyle(
        TableStyle(
            [
                ("FONTNAME", (0, 0), (-1, 0), font_bold),
                ("FONTNAME", (0, 1), (-1, -1), font_regular),
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#198754")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("GRID", (0, 0), (-1, -1), 0.4, colors.lightgrey),
            ]
        )
    )
    story.append(htbl)
    doc.build(story)
    return buffer.getvalue()


def reporting_excel_response(bundle: dict, *, period: str, organization_name: str) -> HttpResponse:
    content = build_reporting_excel(bundle, period=period, organization_name=organization_name)
    response = HttpResponse(
        content,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="kpi_report.xlsx"'
    return response


def reporting_pdf_response(bundle: dict, *, period: str, organization_name: str) -> HttpResponse:
    content = build_reporting_pdf(bundle, period=period, organization_name=organization_name)
    response = HttpResponse(content, content_type="application/pdf")
    response["Content-Disposition"] = 'attachment; filename="kpi_report.pdf"'
    return response
